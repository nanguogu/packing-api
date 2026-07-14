"""Parse the stable labels used by Lichen production-order workbooks."""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook


FIELD_LABELS = {
    "款式类型": "style_type",
    "数量": "quantity_text",
    "公分数": "nominal_size_cm",
    "正面面板": "front_panel",
    "背面面板": "back_panel",
    "字壳围边": "sidewall",
    "底板": "backing",
    "安装脚": "installation_foot",
    "可拆": "detachable",
    "包装": "packaging_type",
    "打箱": "boxing_mode",
    "安装方式": "installation_method",
    "配件选择": "accessories",
    "灯源": "light_source",
    "出线长度": "cable_length",
    "电源": "power_supply",
}


def _nearest_value(row: list[object], start: int) -> object | None:
    for value in row[start + 1:start + 5]:
        if value is not None and str(value).strip():
            return value
    return None


def _first_number(value: object | None) -> float | None:
    if value is None:
        return None
    match = re.search(r"\d+(?:\.\d+)?", str(value))
    return float(match.group()) if match else None


def _quantity(value: object | None) -> int | None:
    number = _first_number(value)
    return int(number) if number is not None else None


def parse_production_sheet(content: bytes, filename: str | None = None) -> dict:
    """Extract labelled production facts without relying on fixed cell addresses."""
    if not content:
        raise ValueError("Production sheet is empty")
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    sheet = workbook.active
    fields: dict[str, object] = {}
    evidence = []
    for row_index, cells in enumerate(sheet.iter_rows(), start=1):
        values = [cell.value for cell in cells]
        for column_index, raw_label in enumerate(values):
            label = str(raw_label).strip() if raw_label is not None else ""
            field = FIELD_LABELS.get(label)
            if not field or field in fields:
                continue
            value = _nearest_value(values, column_index)
            if value is None:
                continue
            fields[field] = value
            evidence.append({
                "field": field,
                "label": label,
                "value": value,
                "cell": f"{cells[column_index].column_letter}{row_index}",
            })

    title = None
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5)):
        for cell in row:
            if isinstance(cell.value, str) and ("单-" in cell.value or "单_" in cell.value):
                title = cell.value.strip()
                break
        if title:
            break

    fields["quantity"] = _quantity(fields.get("quantity_text"))
    fields["nominal_size_cm"] = _first_number(fields.get("nominal_size_cm"))
    return {
        "filename": filename,
        "sheet_name": sheet.title,
        "title": title,
        "fields": fields,
        "evidence": evidence,
    }
